from django.forms import formset_factory, inlineformset_factory
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required, permission_required
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, View
from django.urls import reverse_lazy, reverse
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin, UserPassesTestMixin
from django.db import transaction
from django.db.models import Q, Sum, F, Count, Case, When, Value, IntegerField
from django.utils import timezone
from datetime import timedelta, datetime
from django.db.models.functions import TruncDay, TruncMonth, TruncYear
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect, Http404
from django.template.loader import render_to_string
from django.views.decorators.http import require_http_methods, require_POST
from rest_framework import generics, status, viewsets
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import SessionAuthentication
from io import BytesIO
import csv
from .models import (
    Cliente, Vehiculo, Servicio, Empleado, Reparacion, Tarea,
    TareaHistorial, Agenda, Registro
)
from .forms import (
    ClienteForm, VehiculoForm, ServicioForm, EmpleadoForm,
    ReparacionForm, TareaForm, CitaForm
)
from .serializers import (
    ClienteSerializer, VehiculoSerializer, ServicioSerializer,
    EmpleadoSerializer, ReparacionSerializer, AgendaSerializer, RegistroSerializer
)

User = get_user_model()

# Funciones de ayuda para permisos
def es_jefe(user):
    """Verifica si el usuario es jefe"""
    if not user.is_authenticated:
        return False
    return hasattr(user, 'profile') and hasattr(user.profile, 'es_jefe') and user.profile.es_jefe

def es_encargado(user):
    """Verifica si el usuario es encargado"""
    if not user.is_authenticated:
        return False
    return hasattr(user, 'profile') and hasattr(user.profile, 'es_encargado') and user.profile.es_encargado

def puede_gestionar_empleados(user):
    """Verifica si el usuario puede gestionar empleados"""
    if not user.is_authenticated:
        return False
    return es_jefe(user) or es_encargado(user)

def puede_gestionar_servicios(user):
    """Verifica si el usuario puede gestionar servicios"""
    if not user.is_authenticated:
        return False
    return es_jefe(user) or es_encargado(user)

def es_jefe_o_encargado(user):
    """Verifica si el usuario es jefe o encargado"""
    if not user.is_authenticated:
        return False

    # Si es superusuario, tiene acceso total
    if user.is_superuser:
        return True

    # Verificar si el perfil existe y tiene un empleado relacionado
    if hasattr(user, 'profile') and hasattr(user.profile, 'empleado_relacionado'):
        empleado = user.profile.empleado_relacionado
        if empleado and hasattr(empleado, 'puesto'):
            return empleado.puesto.lower() in ['jefe', 'encargado', 'supervisor', 'administrador']

    return False


def es_mecanico(user):
    """Verifica si el usuario es un mecánico"""
    if not user.is_authenticated:
        return False

    # Verificar si el perfil existe, es empleado y tiene un empleado relacionado
    if hasattr(user, 'profile') and user.profile.es_empleado and hasattr(user.profile, 'empleado_relacionado'):
        empleado = user.profile.empleado_relacionado
        if empleado and hasattr(empleado, 'puesto'):
            return empleado.puesto.lower() in ['mecanico', 'técnico', 'taller']

    return False

# ========== AUTENTICACIÓN Y DASHBOARD BÁSICOS ==========

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, 'Inicio de sesión exitoso.')
            return redirect('inicio')
        messages.error(request, 'Usuario o contraseña incorrectos.')
    return render(request, 'auth/login.html')


@login_required
def logout_view(request):
    logout(request)
    messages.info(request, 'Has cerrado sesión.')
    return redirect('login')


@login_required
def perfil_view(request):
    return render(request, 'auth/perfil.html', {'user': request.user})


@login_required
def listar_tareas(request):
    """
    Vista para listar todas las tareas del usuario.
    """
    # Obtener tareas según el rol del usuario
    if request.user.profile.es_empleado:
        # Si es empleado, mostrar sus tareas asignadas y las que creó
        tareas = Tarea.objects.filter(
            Q(asignada_a=request.user) | Q(creada_por=request.user)
        ).distinct().order_by('fecha_limite', 'prioridad')
    else:
        # Si es jefe o admin, mostrar todas las tareas
        tareas = Tarea.objects.all().order_by('fecha_limite', 'prioridad')

    # Separar tareas por estado
    tareas_por_hacer = tareas.filter(estado='por_hacer')
    tareas_en_progreso = tareas.filter(estado='en_progreso')
    tareas_completadas = tareas.filter(estado='completada')

    context = {
        'tareas_por_hacer': tareas_por_hacer,
        'tareas_en_progreso': tareas_en_progreso,
        'tareas_completadas': tareas_completadas,
    }

    return render(request, 'gestion/tareas/lista_tareas.html', context)

def crear_tarea(request):
    """
    Vista para crear una nueva tarea.
    """
    if request.method == 'POST':
        form = TareaForm(request.POST, user=request.user)
        if form.is_valid():
            tarea = form.save(commit=False)
            tarea.creada_por = request.user
            tarea.save()
            messages.success(request, 'Tarea creada exitosamente.')
            return redirect('lista_tareas')
    else:
        form = TareaForm(user=request.user)

    return render(request, 'gestion/tareas/crear_tarea.html', {'form': form})

def editar_tarea(request, tarea_id):
    """
    Vista para editar una tarea existente.
    """
    tarea = get_object_or_404(Tarea, id=tarea_id)

    # Verificar permisos
    if not (request.user == tarea.creada_por or request.user == tarea.asignada_a or request.user.is_superuser):
        messages.error(request, 'No tienes permiso para editar esta tarea.')
        return redirect('lista_tareas')

    if request.method == 'POST':
        form = TareaForm(request.POST, instance=tarea, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tarea actualizada exitosamente.')
            return redirect('lista_tareas')
    else:
        form = TareaForm(instance=tarea, user=request.user)

    return render(request, 'gestion/tareas/editar_tarea.html', {'form': form, 'tarea': tarea})

@require_POST
def eliminar_tarea(request, tarea_id):
    """
    Vista para eliminar una tarea.
    """
    tarea = get_object_or_404(Tarea, id=tarea_id)

    # Verificar permisos
    if not (request.user == tarea.creada_por or request.user.is_superuser):
        messages.error(request, 'No tienes permiso para eliminar esta tarea.')
        return redirect('lista_tareas')

    tarea.delete()
    messages.success(request, 'Tarea eliminada exitosamente.')
    return redirect('lista_tareas')

import logging
logger = logging.getLogger(__name__)

@require_http_methods(["POST"])
def cambiar_estado_tarea(request, tarea_id, nuevo_estado):
    """
    Vista para cambiar el estado de una tarea mediante AJAX.
    """
    logger.info(f'Cambiando estado de tarea {tarea_id} a {nuevo_estado}')

    try:
        tarea = get_object_or_404(Tarea, id=tarea_id)
        logger.debug(f'Tarea encontrada: {tarea}')

        # Verificar permisos
        if not (request.user == tarea.creada_por or request.user == tarea.asignada_a or request.user.is_superuser):
            error_msg = f'Usuario {request.user} no tiene permiso para modificar la tarea {tarea_id}'
            logger.warning(error_msg)
            return JsonResponse(
                {'success': False, 'error': 'No tienes permiso para modificar esta tarea.'},
                status=403
            )

        # Validar el nuevo estado
        estados_validos = dict(Tarea.ESTADOS_TAREA).keys()
        if nuevo_estado not in estados_validos:
            error_msg = f'Estado {nuevo_estado} no válido. Estados válidos: {estados_validos}'
            logger.warning(error_msg)
            return JsonResponse(
                {'success': False, 'error': 'Estado no válido.'},
                status=400
            )

        # Actualizar el estado
        logger.debug(f'Actualizando tarea {tarea_id} de {tarea.estado} a {nuevo_estado}')
        tarea.estado = nuevo_estado
        tarea.save()

        logger.info(f'Tarea {tarea_id} actualizada exitosamente a {nuevo_estado}')
        return JsonResponse({
            'success': True,
            'nuevo_estado': tarea.get_estado_display(),
            'tarea_id': tarea.id,
            'estado_anterior': tarea.estado,
            'estado_nuevo': nuevo_estado
        })

    except Exception as e:
        logger.error(f'Error al cambiar estado de tarea {tarea_id}: {str(e)}', exc_info=True)
        return JsonResponse(
            {'success': False, 'error': f'Error interno del servidor: {str(e)}'},
            status=500
        )

@login_required
def inicio(request):
    # Redirigir según el rol del usuario
    if es_jefe_o_encargado(request.user):
        return redirect('dashboard_encargado')
    elif es_mecanico(request.user):
        return redirect('dashboard_mecanico')

    # Si no es ni jefe, ni encargado, ni mecánico, mostrar dashboard básico
    total_clientes = Cliente.objects.count()
    total_empleados = Empleado.objects.count()
    total_servicios = Servicio.objects.count()
    total_vehiculos = Vehiculo.objects.count()
    reparaciones_pendientes = Reparacion.objects.filter(estado_reparacion='en_progreso').count()
    # citas_hoy = 0  # Comentado temporalmente hasta que se implemente el modelo Agenda

    context = {
        'total_clientes': total_clientes,
        'total_empleados': total_empleados,
        'total_servicios': total_servicios,
        'total_vehiculos': total_vehiculos,
        'reparaciones_pendientes': reparaciones_pendientes,
        # 'citas_hoy': citas_hoy,  # Comentado temporalmente
    }
    return render(request, 'inicio.html', context)


# Vista temporal para otras rutas aún no implementadas
@login_required
def not_implemented_view(request, *args, **kwargs):
    messages.info(request, 'Funcionalidad en desarrollo.')
    return redirect('inicio')

@login_required
def dashboard_encargado(request):
    """
    Vista del dashboard para el rol de encargado.
    Muestra información relevante para el seguimiento de tareas diarias.
    """
    # Verificar permisos
    if not es_jefe_o_encargado(request.user):
        messages.error(request, 'No tienes permiso para acceder a esta sección.')
        return redirect('inicio')

    # Obtener la fecha de hoy
    hoy = timezone.now().date()

    # Obtener reparaciones en progreso
    reparaciones_en_progreso = Reparacion.objects.filter(
        estado_reparacion='en_progreso'
    ).select_related('vehiculo', 'vehiculo__cliente', 'servicio').order_by('-fecha_ingreso')[:5]

    # Inicializar variables para citas (comentadas temporalmente)
    citas_hoy = []
    proximas_citas = []

    # Obtener tareas para el dashboard
    tareas = Tarea.objects.all()
    tareas_por_hacer = tareas.filter(estado='por_hacer').order_by('-fecha_creacion')[:5]
    tareas_en_progreso = tareas.filter(estado='en_progreso').order_by('-fecha_creacion')[:5]
    tareas_completadas = tareas.filter(estado='completada').order_by('-fecha_creacion')[:5]

    context = {
        'citas_hoy': citas_hoy,
        'reparaciones_en_progreso': reparaciones_en_progreso,
        'proximas_citas': proximas_citas,
        'tareas_por_hacer': tareas_por_hacer,
        'tareas_en_progreso': tareas_en_progreso,
        'tareas_completadas': tareas_completadas,
    }

    return render(request, 'gestion/dashboard_encargado.html', context)


@login_required
def dashboard_mecanico(request):
    """
    Vista del dashboard para el rol de mecánico.
    Muestra información relevante para el trabajo diario del mecánico.
    """
    # Verificar permisos
    if not es_mecanico(request.user):
        messages.error(request, 'No tienes permiso para acceder a esta sección.')
        return redirect('inicio')

    # Obtener fecha actual
    hoy = timezone.now().date()
    mes_actual = timezone.now().month
    año_actual = timezone.now().year

    # Obtener el perfil de empleado del usuario actual
    empleado = None
    if hasattr(request.user, 'profile') and hasattr(request.user.profile, 'empleado_relacionado'):
        empleado = request.user.profile.empleado_relacionado

    # Inicializar variables
    reparaciones_asignadas = Reparacion.objects.none()
    reparaciones_disponibles = Reparacion.objects.none()
    reparaciones_completadas_mes = 0
    reparaciones_en_progreso = 0
    tiempo_promedio_reparacion = None

    if empleado:
        # Reparaciones asignadas activas
        reparaciones_asignadas = Reparacion.objects.filter(
            mecanico_asignado=empleado,
            estado_reparacion__in=['en_progreso', 'pendiente', 'en_espera']
        ).select_related('vehiculo__cliente', 'servicio').order_by('fecha_ingreso')

        # Reparaciones disponibles para tomar
        reparaciones_disponibles = Reparacion.objects.filter(
            estado_reparacion='pendiente',
            mecanico_asignado__isnull=True
        ).select_related('vehiculo__cliente', 'servicio').order_by('fecha_ingreso')[:10]

        # Estadísticas del mes
        reparaciones_completadas_mes = Reparacion.objects.filter(
            mecanico_asignado=empleado,
            estado_reparacion='completada',
            fecha_salida__month=mes_actual,
            fecha_salida__year=año_actual
        ).count()

        reparaciones_en_progreso = reparaciones_asignadas.filter(
            estado_reparacion='en_progreso'
        ).count()

        # Calcular tiempo promedio de reparación (últimas completadas)
        completadas = Reparacion.objects.filter(
            mecanico_asignado=empleado,
            estado_reparacion='completada',
            fecha_salida__isnull=False
        ).order_by('-fecha_salida')[:10]

        if completadas.exists():
            duraciones = [(r.fecha_salida - r.fecha_ingreso).days for r in completadas if r.fecha_salida]
            if duraciones:
                avg_days = sum(duraciones) / len(duraciones)
                tiempo_promedio_reparacion = round(avg_days, 1)

    # Tareas asignadas pendientes
    tareas_asignadas = Tarea.objects.filter(
        asignada_a=request.user,
        estado__in=['por_hacer', 'en_progreso']
    ).order_by('fecha_limite')

    tareas_urgentes = tareas_asignadas.filter(
        fecha_limite__lte=hoy + timezone.timedelta(days=2)
    ).count()

    # Tareas completadas este mes
    tareas_completadas_mes = Tarea.objects.filter(
        asignada_a=request.user,
        estado='completada',
        fecha_actualizacion__month=mes_actual,
        fecha_actualizacion__year=año_actual
    ).count()

    # Tareas recientemente completadas
    tareas_completadas = Tarea.objects.filter(
        asignada_a=request.user,
        estado='completada'
    ).order_by('-fecha_actualizacion')[:5]

    # Próximas citas (placeholder - cuando se implemente el modelo Agenda)
    citas_proximas = []

    context = {
        'titulo': 'Panel del Mecánico',
        'empleado': empleado,
        'hoy': hoy,

        # Reparaciones
        'reparaciones_asignadas': reparaciones_asignadas,
        'reparaciones_disponibles': reparaciones_disponibles,
        'reparaciones_completadas_mes': reparaciones_completadas_mes,
        'reparaciones_en_progreso': reparaciones_en_progreso,
        'tiempo_promedio_reparacion': tiempo_promedio_reparacion,

        # Tareas
        'tareas_asignadas': tareas_asignadas,
        'tareas_urgentes': tareas_urgentes,
        'tareas_completadas': tareas_completadas,
        'tareas_completadas_mes': tareas_completadas_mes,

        # Citas
        'citas_proximas': citas_proximas,
    }

    return render(request, 'gestion/dashboard_mecanico.html', context)


@login_required
def gestionar_reparacion_mecanico(request, reparacion_id):
    """
    Vista para que el mecánico gestione una reparación específica.
    Muestra datos del cliente, permite actualizar datos del vehículo y crear informes.
    """
    # Verificar permisos
    if not es_mecanico(request.user):
        messages.error(request, 'No tienes permiso para acceder a esta sección.')
        return redirect('inicio')

    # Obtener la reparación
    reparacion = get_object_or_404(Reparacion, id=reparacion_id)

    # Obtener el empleado del usuario actual
    empleado = None
    if hasattr(request.user, 'profile') and hasattr(request.user.profile, 'empleado_relacionado'):
        empleado = request.user.profile.empleado_relacionado

    # Si el mecánico está tomando la reparación (no tiene mecánico asignado aún)
    if not reparacion.mecanico_asignado and empleado:
        reparacion.mecanico_asignado = empleado
        if reparacion.estado_reparacion == 'pendiente':
            reparacion.estado_reparacion = 'en_progreso'
        reparacion.save()
        messages.success(request, f'Has tomado la reparación #{reparacion.id}')

    # Verificar que el mecánico asignado sea el usuario actual (o sea admin)
    if reparacion.mecanico_asignado != empleado and not request.user.is_superuser:
        messages.error(request, 'Esta reparación está asignada a otro mecánico.')
        return redirect('dashboard_mecanico')

    # Procesar formulario de actualización
    if request.method == 'POST':
        # Actualizar datos del vehículo extra
        kilometraje = request.POST.get('kilometraje', '').strip()
        nivel_combustible = request.POST.get('nivel_combustible', '').strip()
        observaciones_vehiculo = request.POST.get('observaciones_vehiculo', '').strip()

        # Actualizar condición y estado de reparación
        condicion_vehiculo = request.POST.get('condicion_vehiculo', reparacion.condicion_vehiculo)
        estado_reparacion = request.POST.get('estado_reparacion', reparacion.estado_reparacion)

        # Actualizar informe de reparación
        informe = request.POST.get('informe', '').strip()

        # Construir las notas completas
        notas_completas = []
        if kilometraje:
            notas_completas.append(f"Kilometraje: {kilometraje} km")
        if nivel_combustible:
            notas_completas.append(f"Nivel de combustible: {nivel_combustible}")
        if observaciones_vehiculo:
            notas_completas.append(f"Observaciones del vehículo: {observaciones_vehiculo}")
        if informe:
            notas_completas.append(f"\n--- INFORME DE REPARACIÓN ---\n{informe}")

        # Actualizar la reparación
        reparacion.condicion_vehiculo = condicion_vehiculo
        reparacion.estado_reparacion = estado_reparacion

        # Combinar notas antiguas con nuevas si existen
        if notas_completas:
            nuevas_notas = "\n".join(notas_completas)
            if reparacion.notas:
                reparacion.notas += f"\n\n--- Actualización {timezone.now().strftime('%d/%m/%Y %H:%M')} ---\n{nuevas_notas}"
            else:
                reparacion.notas = nuevas_notas

        # Si se completa la reparación, establecer fecha de salida
        if estado_reparacion == 'completada' and not reparacion.fecha_salida:
            reparacion.fecha_salida = timezone.now()

        reparacion.save()
        messages.success(request, 'Reparación actualizada correctamente.')
        return redirect('gestionar_reparacion_mecanico', reparacion_id=reparacion.id)

    # Extraer información de las notas si existen
    kilometraje_actual = ''
    nivel_combustible_actual = ''
    observaciones_actual = ''

    if reparacion.notas:
        # Intentar extraer info de las notas
        for linea in reparacion.notas.split('\n'):
            if 'Kilometraje:' in linea:
                kilometraje_actual = linea.split('Kilometraje:')[1].strip().replace(' km', '')
            elif 'Nivel de combustible:' in linea:
                nivel_combustible_actual = linea.split('Nivel de combustible:')[1].strip()
            elif 'Observaciones del vehículo:' in linea:
                observaciones_actual = linea.split('Observaciones del vehículo:')[1].strip()

    context = {
        'titulo': f'Gestionar Reparación #{reparacion.id}',
        'reparacion': reparacion,
        'cliente': reparacion.vehiculo.cliente,
        'vehiculo': reparacion.vehiculo,
        'empleado': empleado,
        'kilometraje_actual': kilometraje_actual,
        'nivel_combustible_actual': nivel_combustible_actual,
        'observaciones_actual': observaciones_actual,
        'condicion_opciones': Reparacion.CONDICION_OPCIONES,
        'estado_opciones': Reparacion.ESTADO_REPARACION,
    }

    return render(request, 'gestion/gestionar_reparacion_mecanico.html', context)


@login_required
def buscar_clientes(request):
    q = request.GET.get('q', '').strip()
    clientes = Cliente.objects.all()
    if q:
        clientes = clientes.filter(
            Q(nombre__icontains=q) |
            Q(apellido__icontains=q) |
            Q(correo_electronico__icontains=q) |
            Q(telefono__icontains=q)
        )
    results = []
    for c in clientes[:10]:
        results.append({
            'id': c.id,
            'nombre': c.nombre,
            'apellido': c.apellido,
            'telefono': c.telefono,
            'correo_electronico': c.correo_electronico,
            # Campo alias para compatibilidad con template de vehículos
            'cedula': c.telefono or ''
        })
    # Devolver en dos formatos por compatibilidad con distintos JS en templates
    return JsonResponse({'results': results, 'clientes': results})

# ========== DASHBOARD JEFE ==========

@login_required
def dashboard_jefe(request):
    hoy = timezone.now().date()

    total_clientes = Cliente.objects.count()
    total_vehiculos = Vehiculo.objects.count()
    total_reparaciones = Reparacion.objects.count()
    total_servicios = Servicio.objects.count()
    reparaciones_pendientes = Reparacion.objects.filter(
        estado_reparacion__in=['pendiente', 'en_progreso', 'en_espera', 'revision']
    ).count()
    reparaciones_completadas = Reparacion.objects.filter(estado_reparacion='completada').count()
    # citas_hoy_count = Agenda.objects.filter(fecha=hoy).count()  # Agenda model not implemented yet
    citas_hoy_count = 0  # Placeholder until Agenda model is implemented
    clientes_nuevos_este_mes = Cliente.objects.filter(
        fecha_registro__year=hoy.year,
        fecha_registro__month=hoy.month
    ).count()

    # Vehículos con más reparaciones en últimos 30 días
    desde = hoy - timedelta(days=30)
    vehiculos_frecuentes = Vehiculo.objects.annotate(
        num_reparaciones=Count('reparaciones', filter=Q(reparaciones__fecha_ingreso__gte=desde))
    ).order_by('-num_reparaciones')[:5]

    # Reparaciones por estado (para gráfico)
    estado_map = {
        'pendiente': 'Pendiente',
        'en_progreso': 'En Proceso',
        'en_espera': 'En Espera',
        'revision': 'Para Revisión',
        'completada': 'Completada',
        'cancelada': 'Cancelada',
    }
    rep_estados_qs = Reparacion.objects.values('estado_reparacion').annotate(total=Count('id'))
    reparaciones_por_estado = [
        {
            'estado': estado_map.get(item['estado_reparacion'], item['estado_reparacion']),
            'total': item['total']
        }
        for item in rep_estados_qs
    ]

    # Ingresos mensuales (suma de costo del servicio por mes) - últimos 6 meses
    ingresos_qs = (Reparacion.objects
                   .annotate(m=TruncMonth('fecha_ingreso'))
                   .values('m')
                   .annotate(total=Sum('servicio__costo'))
                   .order_by('m'))
    ingresos_totales_agg = Reparacion.objects.aggregate(total=Sum('servicio__costo'))
    ingresos_totales = float(ingresos_totales_agg['total'] or 0)
    meses_all = [item['m'].strftime('%b %Y') if item['m'] else '' for item in ingresos_qs]
    ingresos_all = [float(item['total']) if item['total'] is not None else 0.0 for item in ingresos_qs]
    meses = meses_all[-6:]
    ingresos = ingresos_all[-6:]
    total_ingresos_mensuales = sum(ingresos) if ingresos else 0.0
    promedio_mensual = (total_ingresos_mensuales / len(ingresos)) if ingresos else None

    # Tiempo promedio de reparación (en días) para completadas
    completadas = Reparacion.objects.filter(fecha_salida__isnull=False)
    if completadas.exists():
        duraciones = [(r.fecha_salida - r.fecha_ingreso).days for r in completadas]
        avg_days = sum(duraciones) / len(duraciones) if duraciones else 0
        # Representar como timedelta
        from datetime import timedelta as _td
        tiempo_promedio = _td(days=int(round(avg_days)))
    else:
        tiempo_promedio = None

    # Listas para secciones
    reparaciones_recientes = Reparacion.objects.select_related('vehiculo', 'servicio').order_by('-fecha_ingreso')[:10]
    # citas_proximas = Agenda.objects.select_related('cliente', 'servicio').filter(fecha__gte=hoy).order_by('fecha', 'hora')[:10]
    citas_proximas = []  # Placeholder until Agenda model is implemented

    # Empleados destacados por registros (últimos 30 días)
    # top_registros = (Registro.objects
    #                  .filter(fecha__gte=desde)
    #                  .values('empleado__id', 'empleado__nombre', 'empleado__puesto')
    #                  .annotate(num_reparaciones=Count('id'))
    #                  .order_by('-num_reparaciones')[:4])
    # empleados_destacados = [
    #     {
    #         'nombre': r['empleado__nombre'],
    #         'puesto': r['empleado__puesto'],
    #         'num_reparaciones': r['num_reparaciones'],
    #     }
    #     for r in top_registros
    # ]
    empleados_destacados = []  # Placeholder until Registro model is implemented

    context = {
        'titulo': 'Panel del Jefe',
        'hoy': timezone.now(),
        'total_clientes': total_clientes,
        'total_vehiculos': total_vehiculos,
        'total_servicios': total_servicios,
        'total_reparaciones': total_reparaciones,
        'reparaciones_pendientes': reparaciones_pendientes,
        'reparaciones_completadas': reparaciones_completadas,
        'citas_hoy_count': citas_hoy_count,
        'clientes_nuevos_este_mes': clientes_nuevos_este_mes,
        'vehiculos_frecuentes': vehiculos_frecuentes,
        'reparaciones_por_estado': reparaciones_por_estado,
        'ingresos_mensuales': bool(ingresos),
        'meses': meses,
        'ingresos': ingresos,
        'ingresos_totales': ingresos_totales,
        'promedio_mensual': promedio_mensual,
        'total_ingresos_mensuales': total_ingresos_mensuales,
        'reparaciones_recientes': reparaciones_recientes,
        'citas_proximas': citas_proximas,
        'empleados_destacados': empleados_destacados,
        'tiempo_promedio': tiempo_promedio,
    }
    return render(request, 'gestion/dashboard_jefe.html', context)

# ========== DASHBOARD REPARACIONES Y CRUD ==========

@login_required
def dashboard_reparaciones(request):
    hoy = timezone.now()

    total_reparaciones = Reparacion.objects.count()
    reparaciones_completadas = Reparacion.objects.filter(estado_reparacion='completada').count()
    reparaciones_en_progreso = Reparacion.objects.filter(estado_reparacion='en_progreso').count()
    reparaciones_pendientes = Reparacion.objects.filter(estado_reparacion='pendiente').count()
    reparaciones_en_espera = Reparacion.objects.filter(estado_reparacion='en_espera').count()
    reparaciones_revision = Reparacion.objects.filter(estado_reparacion='revision').count()
    reparaciones_canceladas = Reparacion.objects.filter(estado_reparacion='cancelada').count()

    # Dict para el gráfico del template
    reparaciones_por_estado = {
        'Completada': reparaciones_completadas,
        'En_Progreso': reparaciones_en_progreso,
        'Pendiente': reparaciones_pendientes,
        'En_Espera': reparaciones_en_espera,
        'Lista_para_Revisión': reparaciones_revision,
        'Cancelada': reparaciones_canceladas,
    }

    # Servicios más solicitados
    servicios_qs = (Reparacion.objects
                    .values('servicio__nombre_servicio')
                    .annotate(total=Count('id'))
                    .order_by('-total')[:5])
    servicios_mas_solicitados = [
        {'nombre': s['servicio__nombre_servicio'], 'total': s['total']}
        for s in servicios_qs
    ]

    ultimas_reparaciones = (Reparacion.objects
                            .select_related('vehiculo', 'servicio', 'vehiculo__cliente')
                            .order_by('-fecha_ingreso')[:10])

    context = {
        'hoy': hoy,
        'total_reparaciones': total_reparaciones,
        'reparaciones_completadas': reparaciones_completadas,
        'reparaciones_en_progreso': reparaciones_en_progreso,
        'reparaciones_pendientes': reparaciones_pendientes,
        'reparaciones_por_estado': reparaciones_por_estado,
        'servicios_mas_solicitados': servicios_mas_solicitados,
        'ultimas_reparaciones': ultimas_reparaciones,
    }
    return render(request, 'gestion/dashboard_reparaciones.html', context)


@login_required
def crear_reparacion(request):
    titulo = 'Nueva Reparación'
    if request.method == 'POST':
        form = ReparacionForm(request.POST)
        if form.is_valid():
            # Crear la reparación pero no guardar aún
            reparacion = form.save(commit=False)
            # Asegurarse de que los campos requeridos tengan valores por defecto si no se proporcionan
            if not reparacion.condicion_vehiculo:
                reparacion.condicion_vehiculo = 'regular'
            if not reparacion.estado_reparacion:
                reparacion.estado_reparacion = 'pendiente'
            # Guardar la reparación con los valores por defecto si es necesario
            reparacion.save()
            messages.success(request, 'Reparación creada correctamente.')
            return redirect('dashboard_reparaciones')
        else:
            # Si el formulario no es válido, agregar los errores a los mensajes
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # Inicializar el formulario con valores por defecto
        form = ReparacionForm(initial={
            'condicion_vehiculo': 'regular',
            'estado_reparacion': 'pendiente'
        })
    return render(request, 'gestion/reparacion_form.html', {'form': form, 'titulo': titulo})


@login_required
def editar_reparacion(request, pk):
    reparacion = get_object_or_404(Reparacion, pk=pk)
    titulo = 'Editar Reparación'
    if request.method == 'POST':
        form = ReparacionForm(request.POST, instance=reparacion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Reparación actualizada correctamente.')
            return redirect('dashboard_reparaciones')
    else:
        form = ReparacionForm(instance=reparacion)
    return render(request, 'gestion/reparacion_form.html', {'form': form, 'titulo': titulo})


@login_required
def eliminar_reparacion(request, pk):
    reparacion = get_object_or_404(Reparacion, pk=pk)
    if request.method == 'POST':
        reparacion.delete()
        messages.success(request, 'Reparación eliminada correctamente.')
        return redirect('dashboard_reparaciones')
    return render(request, 'gestion/reparacion_confirm_delete.html', {'reparacion': reparacion})

# ========== REPORTES ==========

@login_required
def reportes_ingresos(request):
    hoy = timezone.now()
    fecha_desde_str = request.GET.get('fecha_desde')
    fecha_hasta_str = request.GET.get('fecha_hasta')

    reparaciones = Reparacion.objects.all()
    fecha_desde = None
    fecha_hasta = None
    if fecha_desde_str:
        try:
            from datetime import datetime as _dt
            fecha_desde = _dt.strptime(fecha_desde_str, '%Y-%m-%d').date()
            reparaciones = reparaciones.filter(fecha_ingreso__date__gte=fecha_desde)
        except Exception:
            fecha_desde = None
    if fecha_hasta_str:
        try:
            from datetime import datetime as _dt
            fecha_hasta = _dt.strptime(fecha_hasta_str, '%Y-%m-%d').date()
            reparaciones = reparaciones.filter(fecha_ingreso__date__lte=fecha_hasta)
        except Exception:
            fecha_hasta = None

    ingresos_qs = (reparaciones
                   .annotate(m=TruncMonth('fecha_ingreso'))
                   .values('m')
                   .annotate(total=Sum('servicio__costo'), cantidad=Count('id'))
                   .order_by('m'))

    meses_all = [item['m'].strftime('%b %Y') if item['m'] else '' for item in ingresos_qs]
    ingresos_all = [float(item['total']) if item['total'] is not None else 0.0 for item in ingresos_qs]
    cantidades_all = [item['cantidad'] for item in ingresos_qs]

    # Si no hay filtros, mostrar últimos 12 meses; si hay filtros, mostrar todo el rango
    if not fecha_desde_str and not fecha_hasta_str:
        meses = meses_all[-12:]
        ingresos = ingresos_all[-12:]
        cantidades = cantidades_all[-12:]
    else:
        meses = meses_all
        ingresos = ingresos_all
        cantidades = cantidades_all

    ingresos_totales = sum(ingresos) if ingresos else 0.0
    promedio_mensual = (ingresos_totales / len(ingresos)) if ingresos else 0.0

    detalles = [
        {'mes': meses[i], 'total': ingresos[i], 'cantidad': cantidades[i]}
        for i in range(len(meses))
    ]

    context = {
        'titulo': 'Reporte de Ingresos',
        'hoy': hoy,
        'meses': meses,
        'ingresos': ingresos,
        'ingresos_totales': ingresos_totales,
        'promedio_mensual': promedio_mensual,
        'detalles': detalles,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    return render(request, 'gestion/reportes_ingresos.html', context)

# Exportar ingresos a Excel con gráfico (xlsxwriter u openpyxl). Fallback a CSV si ninguna está instalada.
def exportar_ingresos_excel(request):
    fecha_desde_str = request.GET.get('fecha_desde')
    fecha_hasta_str = request.GET.get('fecha_hasta')
    reparaciones = Reparacion.objects.all()

    # Aplicar filtros de fecha si existen
    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
            reparaciones = reparaciones.filter(fecha_ingreso__date__gte=fecha_desde)
        except ValueError:
            pass

    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
            reparaciones = reparaciones.filter(fecha_ingreso__date__lte=fecha_hasta)
        except ValueError:
            pass

    # Obtener datos para el reporte
    ingresos_qs = (reparaciones
                  .annotate(m=TruncMonth('fecha_ingreso'))
                  .values('m')
                  .annotate(
                      total=Sum('servicio__costo'),
                      cantidad=Count('id')
                  )
                  .order_by('m'))

    # Verificar si hay datos para exportar
    if not ingresos_qs:
        return JsonResponse({'error': 'No hay datos para exportar'}, status=400)

    # Procesar datos
    meses = [item['m'].strftime('%b %Y') if item['m'] else '' for item in ingresos_qs]
    ingresos = [float(item['total']) if item['total'] is not None else 0.0 for item in ingresos_qs]
    cantidades = [item['cantidad'] for item in ingresos_qs]

    # Intentar con xlsxwriter (preferido para mejor rendimiento y formato)
    try:
        import xlsxwriter
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Ingresos')

        # Estilos
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4F81BD',  # Azul corporativo
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'font_size': 12
        })

        # Formato para filas pares (gris claro)
        even_row_format = workbook.add_format({
            'bg_color': '#F2F2F2',  # Gris muy claro
            'border': 1,
            'font_size': 11
        })

        # Formato para filas impares (blanco)
        odd_row_format = workbook.add_format({
            'bg_color': '#FFFFFF',  # Blanco
            'border': 1,
            'font_size': 11
        })

        # Formato para moneda
        money_format = workbook.add_format({
            'num_format': '$#,##0.00',
            'border': 1,
            'font_size': 11
        })

        # Formato para celdas de texto
        text_format = workbook.add_format({
            'border': 1,
            'font_size': 11
        })

        # Ancho de columnas
        worksheet.set_column('A:A', 20)  # Mes
        worksheet.set_column('B:B', 25)  # Cantidad
        worksheet.set_column('C:C', 25)  # Ingresos

        # Escribir encabezados
        headers = ['Mes', 'Cantidad de Reparaciones', 'Ingresos Totales']
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)

        # Escribir datos con formato de filas alternadas
        for row_num, (mes, cantidad, ingreso) in enumerate(zip(meses, cantidades, ingresos), start=1):
            # Alternar entre formatos de fila
            row_format = even_row_format if row_num % 2 == 0 else odd_row_format

            # Escribir datos
            worksheet.write(row_num, 0, mes, text_format if row_num % 2 == 1 else text_format)
            worksheet.write_number(row_num, 1, cantidad, row_format)
            worksheet.write_number(row_num, 2, ingreso, money_format if ingreso == 0 else money_format)

        # Crear gráfico
        chart = workbook.add_chart({'type': 'column'})
        last_row = len(meses)
        chart.add_series({
            'name':       'Ingresos',
            'categories': ['Ingresos', 1, 0, last_row, 0],
            'values':     ['Ingresos', 1, 2, last_row, 2],
            'fill':       {'color': '#4F81BD'},  # Mismo azul que el encabezado
            'border':     {'color': '#4F81BD'}
        })

        chart.set_title({'name': 'Ingresos por Mes'})
        chart.set_x_axis({'name': 'Mes'})
        chart.set_y_axis({
            'name': 'Ingresos ($)',
            'num_format': '$#,##0'
        })
        chart.set_legend({'position': 'none'})  # Ocultar leyenda ya que solo hay una serie

        # Insertar gráfico en la hoja
        worksheet.insert_chart('E2', chart, {
            'x_offset': 25,
            'y_offset': 10,
            'x_scale': 1.5,
            'y_scale': 1.5
        })

        # Añadir filtros si se aplicaron fechas
        if fecha_desde_str or fecha_hasta_str:
            filtro_text = "Filtro aplicado: "
            if fecha_desde_str:
                filtro_text += f"Desde {fecha_desde_str}"
            if fecha_hasta_str:
                if fecha_desde_str:
                    filtro_text += " - "
                filtro_text += f"Hasta {fecha_hasta_str}"

            # Escribir texto del filtro
            worksheet.merge_range(f'E{last_row + 3}:H{last_row + 3}', filtro_text, workbook.add_format({
                'italic': True,
                'font_color': '#666666',
                'font_size': 10
            }))

        # Cerrar el libro
        workbook.close()
        output.seek(0)

        # Crear la respuesta
        filename = f"reporte_ingresos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    except ImportError:
        # Si xlsxwriter no está disponible, intentar con openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.chart import BarChart, Reference
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Ingresos"

            # Estilos
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Encabezados
            headers = ['Mes', 'Cantidad de Reparaciones', 'Ingresos Totales']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Datos
            for row_num, (mes, cantidad, ingreso) in enumerate(zip(meses, cantidades, ingresos), start=2):
                # Alternar colores de fila
                if row_num % 2 == 0:
                    fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                else:
                    fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

                # Escribir celdas
                ws.cell(row=row_num, column=1, value=mes).fill = fill
                ws.cell(row=row_num, column=2, value=cantidad).fill = fill
                cell_ingreso = ws.cell(row=row_num, column=3, value=ingreso)
                cell_ingreso.number_format = '$#,##0.00'
                cell_ingreso.fill = fill

                # Aplicar bordes
                for col in range(1, 4):
                    ws.cell(row=row_num, column=col).border = border

            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Crear gráfico
            chart = BarChart()
            chart.title = 'Ingresos por Mes'
            chart.x_axis.title = 'Mes'
            chart.y_axis.title = 'Ingresos ($)'

            data = Reference(ws, min_col=3, min_row=1, max_row=len(meses)+1, max_col=3)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(meses)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            # Personalizar el gráfico
            chart.series[0].graphicalProperties.solidFill = '4F81BD'  # Mismo azul que el encabezado
            chart.series[0].graphicalProperties.line.solidFill = '4F81BD'

            # Añadir el gráfico a la hoja
            ws.add_chart(chart, 'E2')

            # Añadir filtros si se aplicaron fechas
            if fecha_desde_str or fecha_hasta_str:
                filtro_text = "Filtro aplicado: "
                if fecha_desde_str:
                    filtro_text += f"Desde {fecha_desde_str}"
                if fecha_hasta_str:
                    if fecha_desde_str:
                        filtro_text += " - "
                    filtro_text += f"Hasta {fecha_hasta_str}"

                # Escribir texto del filtro
                ws.merge_cells(f'E{len(meses)+3}:H{len(meses)+3}')
                cell_filtro = ws.cell(row=len(meses)+3, column=5, value=filtro_text)
                cell_filtro.font = Font(italic=True, color='666666', size=10)

            # Guardar el libro
            wb.save(output)
            output.seek(0)

            # Crear la respuesta
            filename = f"reporte_ingresos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename={filename}'
            return response

        except ImportError:
            # Si no hay bibliotecas de Excel, usar CSV como último recurso
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename=reporte_ingresos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

            writer = csv.writer(response)
            writer.writerow(['Mes', 'Cantidad Reparaciones', 'Ingresos'])
            for mes, cantidad, ingreso in zip(meses, cantidades, ingresos):
                writer.writerow([mes, cantidad, f'${ingreso:,.2f}'])

            return response

# ========== CLIENTES (CRUD con templates) ==========

@login_required
def clientes_lista(request):
    clientes = Cliente.objects.prefetch_related('vehiculos').order_by('nombre', 'apellido')
    return render(request, 'clientes_lista.html', {'clientes': clientes})


@login_required
@transaction.atomic
def clientes_crear(request):
    VehiculoFormSet = inlineformset_factory(Cliente, Vehiculo, fields=['marca', 'modelo', 'año', 'placa'], extra=1, can_delete=True)
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        formset = VehiculoFormSet(request.POST, prefix='vehiculos')
        if form.is_valid() and formset.is_valid():
            cliente = form.save()
            formset.instance = cliente
            formset.save()
            messages.success(request, 'Cliente creado correctamente.')
            return redirect('clientes-lista')
    else:
        form = ClienteForm()
        formset = VehiculoFormSet(prefix='vehiculos')
    return render(request, 'clientes_form.html', {'form': form, 'formset': formset, 'accion': 'Crear'})


@login_required
@transaction.atomic
def clientes_editar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    VehiculoFormSet = inlineformset_factory(Cliente, Vehiculo, fields=['marca', 'modelo', 'año', 'placa'], extra=0, can_delete=True)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        formset = VehiculoFormSet(request.POST, instance=cliente, prefix='vehiculos')
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, 'Cliente actualizado correctamente.')
            return redirect('clientes-lista')
    else:
        form = ClienteForm(instance=cliente)
        formset = VehiculoFormSet(instance=cliente, prefix='vehiculos')
    return render(request, 'clientes_form.html', {'form': form, 'formset': formset, 'accion': 'Editar'})


@login_required
def clientes_eliminar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        cliente.delete()
        messages.success(request, 'Cliente eliminado correctamente.')
        return redirect('clientes-lista')
    return render(request, 'clientes_confirm_delete.html', {'cliente': cliente})


# ========== EMPLEADOS (CRUD con templates) ==========

@login_required
def empleados_lista(request):
    empleados = Empleado.objects.all().order_by('nombre')
    return render(request, 'empleados_lista.html', {'empleados': empleados})


@login_required
def empleados_crear(request):
    if request.method == 'POST':
        form = EmpleadoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empleado creado correctamente.')
            return redirect('empleados-lista')
    else:
        form = EmpleadoForm()
    return render(request, 'empleados_form.html', {'form': form, 'accion': 'Crear'})


@login_required
def empleados_editar(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    if request.method == 'POST':
        form = EmpleadoForm(request.POST, instance=empleado)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empleado actualizado correctamente.')
            return redirect('empleados-lista')
    else:
        form = EmpleadoForm(instance=empleado)
    return render(request, 'empleados_form.html', {'form': form, 'accion': 'Editar'})


@login_required
def empleados_eliminar(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    if request.method == 'POST':
        empleado.delete()
        messages.success(request, 'Empleado eliminado correctamente.')
        return redirect('empleados-lista')
    return render(request, 'empleados_confirm_delete.html', {'empleado': empleado})


# ========== SERVICIOS (CRUD con templates) ==========

@login_required
def servicios_lista(request):
    servicios = Servicio.objects.all().order_by('nombre_servicio')
    return render(request, 'servicios_lista.html', {'servicios': servicios})


@login_required
def servicios_crear(request):
    if request.method == 'POST':
        form = ServicioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Servicio creado correctamente.')
            return redirect('servicios-lista')
    else:
        form = ServicioForm()
    return render(request, 'servicios_form.html', {'form': form, 'accion': 'Crear'})


@login_required
def servicios_editar(request, pk):
    servicio = get_object_or_404(Servicio, pk=pk)
    if request.method == 'POST':
        form = ServicioForm(request.POST, instance=servicio)
        if form.is_valid():
            form.save()
            messages.success(request, 'Servicio actualizado correctamente.')
            return redirect('servicios-lista')
    else:
        form = ServicioForm(instance=servicio)
    return render(request, 'servicios_form.html', {'form': form, 'accion': 'Editar'})


@login_required
def servicios_eliminar(request, pk):
    servicio = get_object_or_404(Servicio, pk=pk)
    if request.method == 'POST':
        servicio.delete()
        messages.success(request, 'Servicio eliminado correctamente.')
        return redirect('servicios-lista')
    return render(request, 'servicios_confirm_delete.html', {'servicio': servicio})


# ========== VEHÍCULOS (crear/editar/eliminar con template) ==========

@login_required
def vehiculo_agregar(request, cliente_id=None):
    titulo = 'Agregar Vehículo'
    cliente = None
    if cliente_id:
        cliente = get_object_or_404(Cliente, pk=cliente_id)
    if request.method == 'POST':
        form = VehiculoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehículo agregado correctamente.')
            return redirect('vehiculos-lista')
    else:
        form = VehiculoForm(initial={'cliente': cliente.id} if cliente else None)
    return render(request, 'gestion/vehiculo_form.html', {'form': form, 'titulo': titulo, 'cliente': cliente})


@login_required
def vehiculo_editar(request, pk):
    vehiculo = get_object_or_404(Vehiculo, pk=pk)
    titulo = 'Editar Vehículo'
    if request.method == 'POST':
        form = VehiculoForm(request.POST, instance=vehiculo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehículo actualizado correctamente.')
            return redirect('vehiculos-lista')
    else:
        form = VehiculoForm(instance=vehiculo)
    return render(request, 'gestion/vehiculo_form.html', {'form': form, 'titulo': titulo, 'cliente': vehiculo.cliente})


@login_required
def vehiculo_eliminar(request, pk):
    vehiculo = get_object_or_404(Vehiculo, pk=pk)
    if request.method == 'POST':
        vehiculo.delete()
        messages.success(request, 'Vehículo eliminado correctamente.')
        return redirect('vehiculos-lista')
    return render(request, 'vehiculos_confirm_delete.html', {'vehiculo': vehiculo})

# ========== API VIEWS Y VIEWSETS ==========

# Cliente
class ClienteListCreate(generics.ListCreateAPIView):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer


class ClienteRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer


# Empleado
class EmpleadoListCreate(generics.ListCreateAPIView):
    queryset = Empleado.objects.all()
    serializer_class = EmpleadoSerializer


class EmpleadoRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Empleado.objects.all()
    serializer_class = EmpleadoSerializer


# Servicio
class ServicioListCreate(generics.ListCreateAPIView):
    queryset = Servicio.objects.all()
    serializer_class = ServicioSerializer


class ServicioRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Servicio.objects.all()
    serializer_class = ServicioSerializer


# Vehiculo
class VehiculoListCreate(generics.ListCreateAPIView):
    queryset = Vehiculo.objects.all()
    serializer_class = VehiculoSerializer


class VehiculoRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Vehiculo.objects.all()
    serializer_class = VehiculoSerializer


# Reparacion
class ReparacionListCreate(generics.ListCreateAPIView):
    queryset = Reparacion.objects.all()
    serializer_class = ReparacionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Si el usuario es mecánico, solo mostrar sus reparaciones asignadas
        if hasattr(self.request.user, 'profile') and hasattr(self.request.user.profile, 'es_mecanico') and self.request.user.profile.es_mecanico:
            return Reparacion.objects.filter(
                Q(mecanico_asignado__usuario=self.request.user) |
                Q(mecanico_asignado__isnull=True, estado_reparacion='pendiente')
            ).distinct()
        return super().get_queryset()


class ReparacionRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    queryset = Reparacion.objects.all()
    serializer_class = ReparacionSerializer
    permission_classes = [IsAuthenticated]


@login_required
def tomar_reparacion(request, reparacion_id):
    """
    Vista para que un mecánico pueda tomar una reparación.
    """
    # Verificar que el usuario sea un mecánico
    if not hasattr(request.user, 'profile') or not hasattr(request.user.profile, 'es_mecanico') or not request.user.profile.es_mecanico:
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('dashboard')

    # Obtener la reparación
    reparacion = get_object_or_404(Reparacion, id=reparacion_id)

    # Verificar si la reparación ya está asignada a otro mecánico
    if reparacion.mecanico_asignado and reparacion.mecanico_asignado.usuario != request.user:
        messages.warning(request, 'Esta reparación ya ha sido tomada por otro mecánico.')
        return redirect('dashboard_reparaciones')

    # Obtener o crear el perfil de empleado del usuario
    try:
        empleado = request.user.empleado
    except Empleado.DoesNotExist:
        # Si no existe el empleado, crearlo
        empleado = Empleado.objects.create(
            usuario=request.user,
            nombre=request.user.get_full_name() or request.user.username,
            cargo='Mecánico',
            telefono='',
            direccion='',
            fecha_contratacion=timezone.now().date(),
            salario=0
        )

    # Asignar la reparación al mecánico
    reparacion.mecanico_asignado = empleado
    reparacion.estado_reparacion = 'en_progreso'
    reparacion.save()

    messages.success(request, f'Has tomado la reparación de {reparacion.vehiculo}.')
    return redirect('detalle_reparacion', pk=reparacion.id)


@login_required
def listar_reparaciones_disponibles(request):
    """
    Vista para listar las reparaciones disponibles para que un mecánico las tome.
    """
    # Verificar que el usuario sea un mecánico
    if not hasattr(request.user, 'profile') or not hasattr(request.user.profile, 'es_mecanico') or not request.user.profile.es_mecanico:
        messages.error(request, 'No tienes permiso para ver esta página.')
        return redirect('dashboard')

    # Obtener las reparaciones disponibles (sin asignar o asignadas al usuario actual)
    reparaciones = Reparacion.objects.filter(
        Q(mecanico_asignado__isnull=True, estado_reparacion='pendiente') |
        Q(mecanico_asignado__usuario=request.user)
    ).order_by('fecha_ingreso')

    # Obtener las reparaciones asignadas al usuario actual
    mis_reparaciones = Reparacion.objects.filter(
        mecanico_asignado__usuario=request.user,
        estado_reparacion='en_progreso'
    ).order_by('fecha_ingreso')

    return render(request, 'gestion/reparaciones_disponibles.html', {
        'reparaciones': reparaciones,
        'mis_reparaciones': mis_reparaciones,
        'titulo': 'Reparaciones Disponibles'
    })


@login_required
def detalle_reparacion(request, pk):
    """
    Vista para mostrar los detalles de una reparación específica.
    """
    # Obtener la reparación o devolver 404 si no existe
    reparacion = get_object_or_404(Reparacion, pk=pk)

    # Verificar permisos: el usuario debe ser el mecánico asignado o tener permisos de superusuario
    if (not request.user.is_superuser and
        (not hasattr(request.user, 'empleado') or
         reparacion.mecanico_asignado != request.user.empleado)):
        messages.error(request, 'No tienes permiso para ver esta reparación.')
        return redirect('dashboard')

    # Obtener tareas relacionadas con esta reparación
    tareas = Tarea.objects.filter(reparacion=reparacion).order_by('fecha_creacion')

    # Obtener historial de cambios
    historial = TareaHistorial.objects.filter(
        tarea__reparacion=reparacion
    ).select_related('usuario', 'tarea').order_by('-fecha_cambio')

    # Formulario para agregar una nueva tarea
    if request.method == 'POST':
        tarea_form = TareaForm(request.POST, user=request.user)
        if tarea_form.is_valid():
            nueva_tarea = tarea_form.save(commit=False)
            nueva_tarea.reparacion = reparacion
            nueva_tarea.creada_por = request.user
            nueva_tarea.save()

            # Registrar en el historial
            TareaHistorial.objects.create(
                tarea=nueva_tarea,
                usuario=request.user,
                accion='creada',
                descripcion=f'Tarea {nueva_tarea.titulo} creada.'
            )

            messages.success(request, 'Tarea agregada correctamente.')
            return redirect('detalle_reparacion', pk=reparacion.pk)
    else:
        tarea_form = TareaForm(user=request.user)

    return render(request, 'gestion/detalle_reparacion.html', {
        'reparacion': reparacion,
        'tareas': tareas,
        'historial': historial[:10],  # Mostrar solo los 10 registros más recientes
        'tarea_form': tarea_form,
        'titulo': f'Reparación #{reparacion.id} - {reparacion.vehiculo}'
    })


# Agenda (citas) - Comentado temporalmente
# class AgendaViewSet(viewsets.ModelViewSet):
#     queryset = Agenda.objects.all()
#     serializer_class = AgendaSerializer
#
#     def create(self, request, *args, **kwargs):
#         serializer = self.get_serializer(data=request.data)
#         serializer.is_valid(raise_exception=True)
#         try:
#             cita = Agenda().programarCita(
#                 cliente=serializer.validated_data['cliente'],
#                 servicio=serializer.validated_data['servicio'],
#                 fecha=serializer.validated_data['fecha'],
#                 hora=serializer.validated_data['hora']
#             )
#         except ValidationError as e:
#             return Response({'error': e.message}, status=status.HTTP_400_BAD_REQUEST)
#         output_serializer = self.get_serializer(cita)
#         return Response(output_serializer.data, status=status.HTTP_201_CREATED)


# Registro (historial) - Comentado temporalmente
# class RegistroViewSet(viewsets.ModelViewSet):
#     queryset = Registro.objects.all()
#     serializer_class = RegistroSerializer
#
#     def create(self, request, *args, **kwargs):
#         serializer = self.get_serializer(data=request.data)
#         serializer.is_valid(raise_exception=True)
#         try:
#             registro = Registro().crearRegistro(
#                 cliente=serializer.validated_data['cliente'],
#                 empleado=serializer.validated_data['empleado'],
#                 servicio=serializer.validated_data['servicio'],
#                 fecha=serializer.validated_data.get('fecha', None)
#             )
#         except ValidationError as e:
#             return Response({'error': e.message}, status=status.HTTP_400_BAD_REQUEST)
#         output_serializer = self.get_serializer(registro)
#         return Response(output_serializer.data, status=status.HTTP_201_CREATED)


# ========== VISTAS DE CITAS ==========

# @login_required
# def lista_citas(request):
#     """Vista para listar todas las citas"""
#     citas = Agenda.objects.select_related('cliente', 'servicio').order_by('-fecha', '-hora')
#     return render(request, 'gestion/citas/lista_citas.html', {
#         'titulo': 'Lista de Citas',
#         'citas': citas
#     })

# @login_required
# def crear_cita(request):
#     """Vista para crear una nueva cita"""
#     if request.method == 'POST':
#         form = CitaForm(request.POST)
#         if form.is_valid():
#             try:
#                 cita = form.save(commit=False)
#                 # Validar que la fecha no sea pasada
#                 if cita.fecha < timezone.now().date():
#                     messages.error(request, 'No se pueden agendar citas en fechas pasadas.')
#                 # Validar que no haya otra cita a la misma hora
#                 elif Agenda.objects.filter(fecha=cita.fecha, hora=cita.hora).exists():
#                     messages.error(request, 'Ya existe una cita programada para esa fecha y hora.')
#                 else:
#                     cita.save()
#                     messages.success(request, 'Cita creada exitosamente.')
#                     return redirect('lista_citas')
#             except Exception as e:
#                 messages.error(request, f'Error al crear la cita: {str(e)}')
#     else:
#         form = CitaForm()

#     return render(request, 'gestion/citas/crear_cita.html', {
#         'titulo': 'Nueva Cita',
#         'form': form
#     })

# @login_required
# def editar_cita(request, pk):
#     """Vista para editar una cita existente"""
#     cita = get_object_or_404(Agenda, pk=pk)

#     if request.method == 'POST':
#         form = CitaForm(request.POST, instance=cita)
#         if form.is_valid():
#             try:
#                 cita_editada = form.save(commit=False)
#                 # Validar que la fecha no sea pasada
#                 if cita_editada.fecha < timezone.now().date():
#                     messages.error(request, 'No se pueden agendar citas en fechas pasadas.')
#                 # Validar que no haya otra cita a la misma hora (excluyendo la actual)
#                 elif Agenda.objects.filter(fecha=cita_editada.fecha, hora=cita_editada.hora).exclude(pk=pk).exists():
#                     messages.error(request, 'Ya existe otra cita programada para esa fecha y hora.')
#                 else:
#                     cita_editada.save()
#                     messages.success(request, 'Cita actualizada exitosamente.')
#                     return redirect('lista_citas')
#             except Exception as e:
#                 messages.error(request, f'Error al actualizar la cita: {str(e)}')
#     else:
#         form = CitaForm(instance=cita)

#     return render(request, 'gestion/citas/editar_cita.html', {
#         'titulo': 'Editar Cita',
#         'form': form,
#         'cita': cita
#     })

# @login_required
# def detalle_cita(request, pk):
#     """Vista para mostrar los detalles de una cita"""
#     cita = get_object_or_404(Agenda, pk=pk)
#     return render(request, 'gestion/citas/detalle_cita.html', {
#         'titulo': 'Detalle de Cita',
#         'cita': cita
#     })

# @login_required
# def eliminar_cita(request, pk):
#     """Vista para eliminar una cita"""
#     cita = get_object_or_404(Agenda, pk=pk)

#     if request.method == 'POST':
#         try:
#             cita.delete()
#             messages.success(request, 'Cita eliminada exitosamente.')
#         except Exception as e:
#             messages.error(request, f'Error al eliminar la cita: {str(e)}')
#         return redirect('lista_citas')

#     return render(request, 'gestion/citas/eliminar_cita.html', {
#         'titulo': 'Eliminar Cita',
#         'cita': cita
#     })

# def obtener_horas_disponibles(request, fecha):
#     """Vista para obtener las horas disponibles para una fecha dada"""
#     if request.method == 'GET':
#         # Validar que la fecha sea válida
#         try:
#             fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()

#             # Obtener las horas ocupadas para la fecha dada
#             horas_ocupadas = list(Agenda.objects.filter(
#                 fecha=fecha_obj
#             ).values_list('hora', flat=True))

#             # Generar lista de horas disponibles (de 9:00 a 18:00, cada 30 minutos)
#             horas_disponibles = []
#             hora_actual = datetime.strptime('09:00', '%H:%M').time()
#             hora_fin = datetime.strptime('18:00', '%H:%M').time()

#             while hora_actual <= hora_fin:
#                 if hora_actual not in horas_ocupadas:
#                     horas_disponibles.append(hora_actual.strftime('%H:%M'))
#                 # Agregar 30 minutos
#                 hora_actual = (datetime.combine(datetime.today(), hora_actual) + timedelta(minutes=30)).time()

#             return JsonResponse({'horas_disponibles': horas_disponibles})

#         except ValueError:
#             return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)

#     return JsonResponse({'error': 'Método no permitido'}, status=405)


class VehiculoListView(ListView):
    """
    Vista basada en clase para mostrar la lista de vehículos con búsqueda y paginación.
    """
    model = Vehiculo
    template_name = 'gestion/vehiculo_list.html'
    context_object_name = 'vehiculos'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().select_related('cliente')
        busqueda = self.request.GET.get('q', '').strip()

        if busqueda:
            # Búsqueda por placa, marca, modelo o nombre del cliente
            queryset = queryset.filter(
                Q(placa__icontains=busqueda) |
                Q(marca__icontains=busqueda) |
                Q(modelo__icontains=busqueda) |
                Q(cliente__nombre__icontains=busqueda) |
                Q(cliente__apellido__icontains=busqueda)
            )

        return queryset.order_by('-id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Lista de Vehículos'
        context['busqueda'] = self.request.GET.get('q', '')
        return context


class APIAuthenticationMixin:
    """
    Mixin para agregar autenticación a las vistas de API.
    Asegura que solo los usuarios autenticados puedan acceder a las vistas.
    """
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

# Aplicar autenticación a todas las vistas de API
ClienteListCreate.authentication_classes = [SessionAuthentication]
ClienteListCreate.permission_classes = [IsAuthenticated]
ClienteRetrieveUpdateDestroy.authentication_classes = [SessionAuthentication]
ClienteRetrieveUpdateDestroy.permission_classes = [IsAuthenticated]

EmpleadoListCreate.authentication_classes = [SessionAuthentication]
EmpleadoListCreate.permission_classes = [IsAuthenticated]
EmpleadoRetrieveUpdateDestroy.authentication_classes = [SessionAuthentication]
EmpleadoRetrieveUpdateDestroy.permission_classes = [IsAuthenticated]

ServicioListCreate.authentication_classes = [SessionAuthentication]
ServicioListCreate.permission_classes = [IsAuthenticated]
ServicioRetrieveUpdateDestroy.authentication_classes = [SessionAuthentication]
ServicioRetrieveUpdateDestroy.permission_classes = [IsAuthenticated]

VehiculoListCreate.authentication_classes = [SessionAuthentication]
VehiculoListCreate.permission_classes = [IsAuthenticated]
VehiculoRetrieveUpdateDestroy.authentication_classes = [SessionAuthentication]
VehiculoRetrieveUpdateDestroy.permission_classes = [IsAuthenticated]

ReparacionListCreate.authentication_classes = [SessionAuthentication]
ReparacionListCreate.permission_classes = [IsAuthenticated]
ReparacionRetrieveUpdateDestroy.authentication_classes = [SessionAuthentication]
ReparacionRetrieveUpdateDestroy.permission_classes = [IsAuthenticated]

# Comentado temporalmente hasta que se implementen estos modelos
# AgendaViewSet.authentication_classes = [SessionAuthentication]
# AgendaViewSet.permission_classes = [IsAuthenticated]
# RegistroViewSet.authentication_classes = [SessionAuthentication]
# RegistroViewSet.permission_classes = [IsAuthenticated]
